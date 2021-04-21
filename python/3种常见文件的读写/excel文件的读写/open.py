# 安装第三方插件
#//! pip3 install openpyxl

from openpyxl import Workbook  # 引入工作簿
wb = Workbook()   # 创建一个工作簿
ws = wb.active      # 工作表
ws.title = 'Test'   # 将工作表名 'Sheet' 改名为Test
print(ws.title)     # 结果为：Test

ws['A1']= 100  # 在A1单元格写入100，在Python中，excel是从A1开始的，即第1行第1列开始的
ws.cell(row=2, column=2,value=222) # 在第2行第2列的位置处写入222



ws2 = wb.create_sheet('Python') # 创建第二个工作表Python
print(ws2.title)    # 结果为：Python
print(wb.sheetnames) # 结果为：['Test', 'Python']
wb.save('test.xlsx') # 报错，如果文件不存在会自动创建










